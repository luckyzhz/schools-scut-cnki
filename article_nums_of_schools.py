#! python3
# article_nums_of_schools.py - 根据已有数据库，统计每个学院的论文数量

import json
import openpyxl

filename_articles = 'articles-2021.10.23.json'      # 论文数据库
filename_schools = 'school_abbrs_and_names.json'    # 学院缩写和名称数据库

with open(filename_articles, encoding="utf-8") as f:
    # 取得文章 list
    articles = json.load(f)

with open(filename_schools, encoding="utf-8") as f:
    # 取得学院 dictionary
    schools = json.load(f)

articles_of_schools = {}    # 用于存放每个学院的论文数
for school_abbr, school_name in schools.items():
    # 初始化每个学院的论文数为 0
    articles_of_schools[school_abbr] = 0

for article in articles:
    schools_in_one_article = set()  # 一篇文章中，所有学院的集合

    for author in article['authors']:
        for school_abbr in schools.keys():
            # 取得学院教师名录
            filename = f'scut_schools/{school_abbr}.json'
            with open(filename, encoding='utf-8') as f:
                teachers = json.load(f)

            # 如果作者在某个学院，则添加该学院到这篇文章的学院集合
            if author in teachers:
                schools_in_one_article.add(school_abbr)

    # 遍历这篇文章的学院集合，对应学院论文数自增 1
    for school in schools_in_one_article:
        articles_of_schools[school] += 1

# 将数据存储到 xlsx 文件
filename_xlsx = 'article_nums_of_schools.xlsx'
wb = openpyxl.Workbook()    # 新建空白工作簿
sheet = wb.active           # 工作表

i = 1
for school_abbr in articles_of_schools.keys():
    sheet.cell(row=i, column=1).value = schools[school_abbr]
    sheet.cell(row=i, column=2).value = articles_of_schools[school_abbr]
    i += 1

wb.save(filename_xlsx)
